# -*- coding: utf-8 -*-
"""
poller.py — odpytuje modele AI z WŁĄCZONYM WYSZUKIWANIEM INTERNETOWYM
(natywne narzędzia search każdego dostawcy), zapisuje odpowiedzi do plików .txt.

Cel badawczy: odpowiedzi możliwie zbliżone do tego, co widzi użytkownik
przeglądarkowej wersji czatu (warstwa model + wyszukiwarka).

Pytania: questions.xlsx (kolumny: id, pytanie) ma pierwszeństwo; inaczej questions.yaml.
Uruchomienie:  python poller.py            (klucze w zmiennych środowiskowych)
Tryb testowy:  python poller.py --mock     (bez API)
"""
import os, sys, time, json, pathlib, datetime
import requests, yaml

MOCK = "--mock" in sys.argv
TIMEOUT = 180


# ---------------- DOSTAWCY (każdy z natywnym wyszukiwaniem) ----------------

def openai_search(model, pytanie):
    """OpenAI Responses API + narzędzie web_search."""
    r = requests.post(
        "https://api.openai.com/v1/responses",
        headers={"Authorization": f"Bearer {os.environ['OPENAI_API_KEY']}"},
        json={"model": model, "input": pytanie, "tools": [{"type": "web_search"}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    d = r.json()
    if d.get("output_text"):
        return d["output_text"]
    czesci = []
    for blok in d.get("output", []):
        for c in blok.get("content", []) or []:
            if c.get("type") == "output_text":
                czesci.append(c.get("text", ""))
    return "\n".join(czesci)


def anthropic_search(model, pytanie):
    """Anthropic Messages API + narzędzie web_search_20250305."""
    r = requests.post(
        "https://api.anthropic.com/v1/messages",
        headers={"x-api-key": os.environ["ANTHROPIC_API_KEY"],
                 "anthropic-version": "2023-06-01"},
        json={"model": model, "max_tokens": 2048,
              "messages": [{"role": "user", "content": pytanie}],
              "tools": [{"type": "web_search_20250305", "name": "web_search",
                         "max_uses": 5}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    return "\n".join(b.get("text", "") for b in r.json()["content"]
                     if b.get("type") == "text")


def gemini_search(model, pytanie):
    """Gemini generateContent + Grounding with Google Search."""
    r = requests.post(
        f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent",
        headers={"x-goog-api-key": os.environ["GEMINI_API_KEY"]},
        json={"contents": [{"parts": [{"text": pytanie}]}],
              "tools": [{"google_search": {}}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    czesci = r.json()["candidates"][0]["content"]["parts"]
    return "\n".join(p.get("text", "") for p in czesci)


def perplexity_search(model, pytanie):
    """Perplexity sonar — wyszukiwanie natywne, zawsze włączone."""
    r = requests.post(
        "https://api.perplexity.ai/chat/completions",
        headers={"Authorization": f"Bearer {os.environ['PERPLEXITY_API_KEY']}"},
        json={"model": model,
              "messages": [{"role": "user", "content": pytanie}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    d = r.json()
    tekst = d["choices"][0]["message"]["content"]
    zrodla = d.get("citations") or d.get("search_results") or []
    if zrodla:
        tekst += "\n\n[Źródła]\n" + "\n".join(
            z if isinstance(z, str) else z.get("url", "") for z in zrodla)
    return tekst


def xai_search(model, pytanie):
    """xAI Responses API + narzędzie web_search (Agent Tools, następca Live Search)."""
    r = requests.post(
        "https://api.x.ai/v1/responses",
        headers={"Authorization": f"Bearer {os.environ['XAI_API_KEY']}"},
        json={"model": model, "input": pytanie, "tools": [{"type": "web_search"}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    d = r.json()
    if d.get("output_text"):
        return d["output_text"]
    czesci = []
    for blok in d.get("output", []):
        for c in blok.get("content", []) or []:
            if c.get("type") == "output_text":
                czesci.append(c.get("text", ""))
    return "\n".join(czesci)


def openrouter_online(model, pytanie):
    """PRZYBLIŻENIE dla modeli bez oficjalnego API z wyszukiwaniem
    (Meta AI, DeepSeek, Mistral): OpenRouter + wtyczka web (sufiks :online).
    To NIE jest tożsame z wersją przeglądarkową — oznaczamy w metryce."""
    r = requests.post(
        "https://openrouter.ai/api/v1/chat/completions",
        headers={"Authorization": f"Bearer {os.environ['OPENROUTER_API_KEY']}"},
        json={"model": model + ":online",
              "messages": [{"role": "user", "content": pytanie}]},
        timeout=TIMEOUT)
    r.raise_for_status()
    return r.json()["choices"][0]["message"]["content"]


DOSTAWCY = {
    "openai": (openai_search, "natywne wyszukiwanie (web_search)"),
    "anthropic": (anthropic_search, "natywne wyszukiwanie (web_search_20250305)"),
    "gemini": (gemini_search, "natywne wyszukiwanie (Google Search grounding)"),
    "perplexity": (perplexity_search, "wyszukiwanie natywne (sonar)"),
    "xai": (xai_search, "natywne wyszukiwanie (Agent Tools web_search)"),
    "openrouter_online": (openrouter_online,
                          "PRZYBLIŻENIE: OpenRouter + wtyczka web, nie wersja przeglądarkowa"),
}


# ---------------- PYTANIA ----------------

def wczytaj_pytania():
    if pathlib.Path("questions.xlsx").exists():
        from openpyxl import load_workbook
        ws = load_workbook("questions.xlsx").active
        pyt = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and row[1]:
                pyt.append({"id": str(row[0]).strip(), "tekst": str(row[1]).strip()})
        return pyt
    cfg = yaml.safe_load(open("questions.yaml", encoding="utf-8"))
    return cfg["pytania"]


# ---------------- GŁÓWNA PĘTLA ----------------

def main():
    cfg = yaml.safe_load(open("questions.yaml", encoding="utf-8"))
    pytania = wczytaj_pytania()
    teraz = datetime.datetime.now(datetime.timezone.utc)
    katalog = pathlib.Path("answers") / teraz.strftime("%Y-%m-%d_%H-%M")
    katalog.mkdir(parents=True, exist_ok=True)

    bledy, pominiete = [], []
    for pyt in pytania:
        for wpis in cfg["modele"]:
            dostawca, model, etykieta = wpis["dostawca"], wpis["model"], wpis["etykieta"]
            fn, tryb = DOSTAWCY[dostawca]
            for powt in range(1, int(cfg.get("powtorzenia", 1)) + 1):
                plik = katalog / f"{pyt['id']}__{etykieta}__p{powt}.txt"
                if MOCK:
                    odp, status = f"[TEST] {etykieta}: {pyt['tekst'][:60]}", "OK (mock)"
                else:
                    try:
                        odp, status = fn(model, pyt["tekst"]), "OK"
                    except KeyError as e:
                        odp, status = f"[POMINIĘTO] brak klucza: {e}", "POMINIĘTO"
                        pominiete.append(etykieta)
                    except Exception as e:
                        odp, status = f"[BŁĄD] {e}", "BŁĄD"
                        bledy.append(str(plik))
                plik.write_text(
                    f"Pytanie: {pyt['tekst']}\nId pytania: {pyt['id']}\n"
                    f"Model: {etykieta} ({dostawca}/{model})\n"
                    f"Tryb wyszukiwania: {tryb}\n"
                    f"Data (UTC): {teraz.isoformat(timespec='seconds')}\n"
                    f"Powtórzenie: {powt}\nStatus: {status}\n"
                    + "=" * 60 + "\n\n" + odp,
                    encoding="utf-8")
                print("zapisano:", plik, "|", status)
                time.sleep(1)

    (katalog / "_przebieg.json").write_text(json.dumps({
        "start_utc": teraz.isoformat(timespec="seconds"),
        "pytan": len(pytania), "modeli": len(cfg["modele"]),
        "powtorzen": cfg.get("powtorzenia", 1),
        "bledy": bledy, "pominiete_brak_klucza": sorted(set(pominiete)),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    print("GOTOWE. Błędów:", len(bledy), "| pominiętych:", len(set(pominiete)))


if __name__ == "__main__":
    main()
